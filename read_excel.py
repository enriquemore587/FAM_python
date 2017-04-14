from openpyxl import load_workbook

wb = load_workbook('libro1.xlsx')


ws = wb['Sheet1']

celdas = ws['C3:E10']


def matrix_xl(cells):
    mat = []
    for row_xl in cells:
        row = []
        for cell in row_xl:
            row.append(cell.value)
        mat.append(row)
    return mat

print matrix_xl(celdas)